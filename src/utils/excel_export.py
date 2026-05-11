"""Exportação Excel estilizada — padrão NTC Grupo Náutica Refrigeração."""

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
_NAVY     = "14283C"
_WHITE    = "FFFFFF"
_GOLD     = "BFA168"
_ALT      = "F5F7FA"   # linha alternada
_TOTAL_BG = "E7EBF0"   # rodapé de totais
_GREEN    = "1F7A3A"
_RED      = "B5322B"

# ── Formatos de número (códigos Excel) ────────────────────────────────────────
_FMT = {
    "brl":       '"R$ "#,##0.00',
    "pct":       "0.0%",
    "int":       "#,##0",
    "delta_brl": '"R$ "#,##0.00;[Red]"R$ "-#,##0.00',
    "delta_pp":  '[Blue]+0.0%;[Red]-0.0%;"-"',
    "text":      None,
}

_THIN   = Side(style="thin",   color="E4E4E4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TOTAL_FMTS = {"brl", "int", "delta_brl"}   # usa SUM
_AVG_FMTS   = {"pct"}                        # usa AVERAGE se contiver "margem" na key


def to_excel_styled(
    df: pd.DataFrame,
    cols: list[dict],
    sheet_name: str = "Dados",
    show_totals: bool = True,
    freeze_cols: int = 0,
) -> bytes:
    """
    Gera Excel estilizado padrão NTC.

    cols — lista de dicts:
        key        str  — nome da coluna no DataFrame
        label      str  — cabeçalho exibido
        fmt        str  — "brl" | "pct" | "int" | "delta_brl" | "delta_pp" | "text"
        width      int  — largura da coluna (opcional; auto-calculada se omitido)
        total      bool | "sum" | "avg" | None  — comportamento na linha de totais
                   (padrão: auto com base no fmt)

    freeze_cols — número de colunas à esquerda a congelar (além da linha 1)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Apenas colunas que existem no df
    valid = [c for c in cols if c["key"] in df.columns]
    keys  = [c["key"] for c in valid]
    data  = df[keys].copy()
    nrows = len(data)
    ncols = len(valid)

    # ── Estilos base ──────────────────────────────────────────────────────────
    hdr_fill  = PatternFill(fgColor=_NAVY, fill_type="solid")
    hdr_font  = Font(color=_WHITE, bold=True, name="Calibri", size=10)
    alt_fill  = PatternFill(fgColor=_ALT,  fill_type="solid")
    tot_fill  = PatternFill(fgColor=_TOTAL_BG, fill_type="solid")
    tot_font  = Font(bold=True, name="Calibri", size=10, color=_NAVY)
    data_font = Font(name="Calibri", size=10)

    # ── Linha 1: cabeçalho ────────────────────────────────────────────────────
    for ci, col in enumerate(valid, 1):
        cell = ws.cell(row=1, column=ci, value=col["label"])
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER
    ws.row_dimensions[1].height = 24

    # ── Linhas de dados ───────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(data.iterrows(), 2):
        is_alt = (ri % 2 == 0)
        for ci, col in enumerate(valid, 1):
            raw = row[col["key"]]

            # Manter numérico sempre que possível
            if isinstance(raw, str):
                try:
                    raw = float(raw.replace("R$", "").replace("%", "")
                                .replace(".", "").replace(",", ".").strip())
                except (ValueError, AttributeError):
                    pass

            cell = ws.cell(row=ri, column=ci, value=raw)
            cell.font   = data_font
            cell.border = _BORDER
            if is_alt:
                cell.fill = alt_fill

            fmt_code = _FMT.get(col.get("fmt", "text"))
            if fmt_code:
                cell.number_format = fmt_code

            is_numeric_fmt = col.get("fmt", "text") in _FMT and col.get("fmt") != "text"
            cell.alignment = Alignment(
                horizontal="right" if is_numeric_fmt else "left",
                vertical="center",
            )

        ws.row_dimensions[ri].height = 18

    # ── Linha de totais ───────────────────────────────────────────────────────
    if show_totals and nrows > 0:
        tr = nrows + 2
        first_label = True
        for ci, col in enumerate(valid, 1):
            fmt_type  = col.get("fmt", "text")
            col_ltr   = get_column_letter(ci)
            total_cfg = col.get("total")   # None = auto, False = skip, "sum"/"avg" = force

            cell = ws.cell(row=tr, column=ci)
            cell.fill   = tot_fill
            cell.font   = tot_font
            cell.border = _BORDER

            if total_cfg is False:
                cell.value     = ""
                cell.alignment = Alignment(horizontal="left", vertical="center")

            elif fmt_type in _TOTAL_FMTS or total_cfg == "sum":
                cell.value        = f"=SUM({col_ltr}2:{col_ltr}{tr-1})"
                cell.number_format = _FMT.get(fmt_type, "")
                cell.alignment    = Alignment(horizontal="right", vertical="center")

            elif fmt_type in _AVG_FMTS and ("margem" in col["key"].lower() or total_cfg == "avg"):
                cell.value        = f"=AVERAGE({col_ltr}2:{col_ltr}{tr-1})"
                cell.number_format = _FMT["pct"]
                cell.alignment    = Alignment(horizontal="right", vertical="center")

            else:
                if first_label and fmt_type == "text":
                    cell.value = "TOTAL"
                    first_label = False
                else:
                    cell.value = ""
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[tr].height = 22

    # ── Congelar painéis ──────────────────────────────────────────────────────
    freeze_col_ltr = get_column_letter(freeze_cols + 1) if freeze_cols > 0 else "A"
    ws.freeze_panes = f"{freeze_col_ltr}2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    # ── Formatação condicional — colunas de margem ────────────────────────────
    if nrows > 0:
        for ci, col in enumerate(valid, 1):
            if col.get("fmt") == "pct" and "margem" in col["key"].lower():
                col_ltr = get_column_letter(ci)
                rng     = f"{col_ltr}2:{col_ltr}{nrows + 1}"
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="num", start_value=0.0,  start_color="B5322B",
                    mid_type="num",   mid_value=0.15,   mid_color="FFEB99",
                    end_type="num",   end_value=0.35,   end_color="1F7A3A",
                ))

    # ── Larguras das colunas ──────────────────────────────────────────────────
    for ci, col in enumerate(valid, 1):
        col_ltr  = get_column_letter(ci)
        if "width" in col:
            ws.column_dimensions[col_ltr].width = col["width"]
        else:
            fmt_type   = col.get("fmt", "text")
            header_len = len(col["label"])
            if fmt_type in ("brl", "delta_brl"):
                w = max(header_len, 18)
            elif fmt_type in ("pct", "delta_pp"):
                w = max(header_len, 10)
            elif fmt_type == "int":
                w = max(header_len, 10)
            else:
                sample = data[col["key"]].astype(str).str.len()
                w = max(header_len, min(int(sample.max() * 0.85) if not sample.empty else 10, 42))
            ws.column_dimensions[col_ltr].width = w + 2

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
