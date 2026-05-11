"""Testes unitários — src/utils/excel_export.py"""

import math
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook


def _load_wb(data: bytes):
    return load_workbook(BytesIO(data))


def _basic_df():
    return pd.DataFrame({
        "sku":           ["SK-001", "SK-002", "SK-003"],
        "produto":       ["Produto Alpha", "Produto Beta", "Produto Gamma"],
        "receita_total": [1000.0, 500.0, 200.0],
        "custo_produto": [300.0, 150.0, 80.0],
        "margem":        [0.25, 0.15, 0.10],
    })


def _basic_cols():
    return [
        {"key": "sku",           "label": "SKU",                "fmt": "text",  "width": 13},
        {"key": "produto",       "label": "Produto",            "fmt": "text",  "width": 40},
        {"key": "receita_total", "label": "Receita Bruta (R$)", "fmt": "brl"},
        {"key": "custo_produto", "label": "Custo (R$)",         "fmt": "brl"},
        {"key": "margem",        "label": "Margem %",           "fmt": "pct"},
    ]


from src.utils.excel_export import to_excel_styled


class TestToExcelStyled:

    def test_retorna_bytes(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_abre_como_xlsx_valido(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        wb = _load_wb(data)
        assert wb is not None

    def test_nome_sheet_correto(self):
        data = to_excel_styled(_basic_df(), _basic_cols(), sheet_name="MeuSheet")
        wb = _load_wb(data)
        assert "MeuSheet" in wb.sheetnames

    def test_nome_sheet_padrao(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        wb = _load_wb(data)
        assert "Dados" in wb.sheetnames

    def test_cabecalho_na_linha_1(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        assert ws.cell(1, 1).value == "SKU"
        assert ws.cell(1, 2).value == "Produto"
        assert ws.cell(1, 3).value == "Receita Bruta (R$)"

    def test_numero_de_linhas_correto(self):
        df = _basic_df()
        data = to_excel_styled(df, _basic_cols())
        ws = _load_wb(data).active
        # 1 header + 3 data + 1 total = 5 rows max
        data_rows = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)
                     if ws.cell(r, 1).value not in (None, "TOTAL")]
        assert len(data_rows) == len(df)

    def test_dados_inseridos_corretamente(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        skus = [ws.cell(r, 1).value for r in range(2, 5)]
        assert "SK-001" in skus
        assert "SK-002" in skus
        assert "SK-003" in skus

    def test_linha_total_presente(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        last_row = ws.max_row
        total_label = ws.cell(last_row, 1).value
        # Primeira coluna texto → "TOTAL" ou soma
        assert total_label == "TOTAL" or ws.cell(last_row, 3).value is not None

    def test_sem_total_quando_show_totals_false(self):
        df = _basic_df()
        data = to_excel_styled(df, _basic_cols(), show_totals=False)
        ws = _load_wb(data).active
        # Com show_totals=False não há linha TOTAL
        last_cell = ws.cell(ws.max_row, 1).value
        assert last_cell != "TOTAL"
        assert ws.max_row == len(df) + 1  # header + data only

    def test_freeze_panes_ativo(self):
        data = to_excel_styled(_basic_df(), _basic_cols(), freeze_cols=2)
        ws = _load_wb(data).active
        # Com freeze_cols=2, freeze em C2
        assert ws.freeze_panes == "C2"

    def test_freeze_sem_cols_congela_linha_1(self):
        data = to_excel_styled(_basic_df(), _basic_cols(), freeze_cols=0)
        ws = _load_wb(data).active
        assert ws.freeze_panes == "A2"

    def test_auto_filter_aplicado(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith("A1")

    def test_formato_brl_aplicado(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        # Coluna 3 = receita_total = BRL format
        fmt = ws.cell(2, 3).number_format
        assert "R$" in fmt or "#,##0" in fmt

    def test_formato_pct_aplicado(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        # Coluna 5 = margem = pct format
        fmt = ws.cell(2, 5).number_format
        assert "%" in fmt

    def test_coluna_inexistente_no_df_ignorada(self):
        cols_com_extra = _basic_cols() + [
            {"key": "col_inexistente", "label": "Nao Existe", "fmt": "brl"}
        ]
        data = to_excel_styled(_basic_df(), cols_com_extra)
        wb = _load_wb(data)
        assert wb is not None  # não levantou erro

    def test_df_vazio_nao_levanta_erro(self):
        df_vazio = _basic_df().head(0)
        data = to_excel_styled(df_vazio, _basic_cols())
        assert isinstance(data, bytes)

    def test_largura_coluna_customizada(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        # coluna A (SKU) deve ter width=13 conforme spec
        assert ws.column_dimensions["A"].width == 13

    def test_soma_total_brl(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        # Linha de total da coluna receita_total (col 3) deve ser uma fórmula SUM
        total_cell = ws.cell(ws.max_row, 3).value
        assert str(total_cell).startswith("=SUM") or isinstance(total_cell, (int, float))

    def test_total_false_para_coluna_especifica(self):
        cols = _basic_cols()
        cols[0]["total"] = False  # SKU não deve ter total
        data = to_excel_styled(_basic_df(), cols)
        ws = _load_wb(data).active
        # Primeira coluna da linha total deve ser vazia ou "TOTAL"
        val = ws.cell(ws.max_row, 1).value
        assert val in (None, "", "TOTAL")

    def test_formatacao_condicional_em_margem(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        # Deve haver formatação condicional na coluna de margem
        assert len(ws.conditional_formatting._cf_rules) > 0

    def test_cabecalho_cor_navy(self):
        data = to_excel_styled(_basic_df(), _basic_cols())
        ws = _load_wb(data).active
        header_fill = ws.cell(1, 1).fill
        assert header_fill.fgColor.rgb.upper().endswith("14283C")

    def test_sheet_name_truncado_em_31_chars(self):
        nome_longo = "NomeDeAbaExtraMuitoLongoDeMaisParaExcel"
        data = to_excel_styled(_basic_df(), _basic_cols(), sheet_name=nome_longo)
        wb = _load_wb(data)
        # Excel trunca em 31 caracteres
        assert any(len(s) <= 31 for s in wb.sheetnames)

    def test_formatos_delta_brl_e_delta_pp(self):
        df = pd.DataFrame({
            "delta_receita": [100.0, -50.0],
            "delta_margem":  [0.05, -0.02],
        })
        cols = [
            {"key": "delta_receita", "label": "Variacao R$",  "fmt": "delta_brl"},
            {"key": "delta_margem",  "label": "Variacao p.p.", "fmt": "delta_pp"},
        ]
        data = to_excel_styled(df, cols)
        ws = _load_wb(data).active
        fmt_brl = ws.cell(2, 1).number_format
        fmt_pp  = ws.cell(2, 2).number_format
        assert "R$" in fmt_brl or "#" in fmt_brl
        assert "%" in fmt_pp or "0" in fmt_pp

    def test_largura_auto_para_coluna_texto_sem_width(self):
        df = pd.DataFrame({"descricao": ["Produto muito longo mesmo", "Curto", "Médio comprimento"]})
        cols = [{"key": "descricao", "label": "Descrição", "fmt": "text"}]  # sem width
        data = to_excel_styled(df, cols)
        ws = _load_wb(data).active
        # A largura deve ser calculada automaticamente (> 0)
        width = ws.column_dimensions["A"].width
        assert width > 0

    def test_formato_int(self):
        df = pd.DataFrame({"qtd": [10, 20, 30]})
        cols = [{"key": "qtd", "label": "Qtd", "fmt": "int"}]
        data = to_excel_styled(df, cols)
        ws = _load_wb(data).active
        fmt = ws.cell(2, 1).number_format
        assert "#,##0" in fmt or fmt in ("#,##0", "General")
